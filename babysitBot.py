#!/usr/bin/env python
# -*- coding: utf-8 -*-
## ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ ##
##                                                                                   ##
##  ----  ----  ----     BOT TELEGRAM SUIVI GARDE A DOMICILE     ----  ----  ----  ##
##                                                                                   ##
## ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

## programme pour suivre via Telegram les horaires de babysit et calculer la rémunération associée.

## ~~~~~~~~~~~~~~~~~~~~~~~~~~        PARAMETRES         ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# modules complémentaires
import os
import sys
import sqlite3
from re import compile as reCompile
from datetime import datetime as dt
from locale import setlocale, LC_ALL
from traceback import format_exc
from hashlib import md5
# modules complémentaires externes
from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler
from telegram.ext import ConversationHandler, MessageHandler, Filters, MessageFilter
from openpyxl import Workbook


# les erreurs critiques
class Exit(Exception):
    """gestion des erreurs pendant la manipulation de la base de donnée."""

    pass

# mise du programme en français pour les affichage de strftime
setlocale(LC_ALL, 'fr_FR.utf8')

# dossiers racine du projet
BASEPATH = os.path.dirname(os.path.realpath(sys.argv[0]))
# chemin vers la base de donnée
BDD_PATH = os.path.join(BASEPATH, "data.db")
# table de la base de donnée
BDD_TABLE = "journees"

# configuration du .env
REGEX_TOKEN = reCompile("token=[0-9]{8,10}:[a-zA-Z0-9_-]{35}")

# les demandes pour la création d'un nouvel enregistrement
DATE, LIEU, H_DEBUT, H_FIN, TRAJET, COMMENTAIRE = range(6)
# le buffer pour enregistrer les infos
TO_SAVE = []
# la colonne maxi dans laquelle on va écrire les données sur excel (lié au modèle de suivi des fiches de paies)
MAX_COL = 5


## ~~~~~~~~~~~~~~~~~~~~~~~~~~      GESTION DU SQL       ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# la classe qui va contenir la base de donnée
class obj_bdd():
    # fonction d'initialisation et de fermeture de la connection
    def __init__(self, FULLPATH, tableName):
        try:
            # si la base de donnée n'existe pas
            if not os.path.isfile(FULLPATH):
                with open(FULLPATH, "w+") as f:
                    pass
            # curseur et connection de la base de donnée
            self._conn = sqlite3.connect(FULLPATH)
            self._cursor = self._conn.cursor()
            # vérification du nom de la table
            self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            self.tableName = tableName
            listeTable = [k[0] for k in self.cursor.fetchall()]
            # si la table n'existe pas, on la crée
            if self.tableName not in listeTable:
                self.cursor.execute(f"CREATE TABLE IF NOT EXISTS '{BDD_TABLE}' ('id' TEXT PRIMARY KEY, 'username' TEXT, 'date' TEXT, 'heure_debut' TEXT, 'heure_fin' TEXT, 'trajet' TEXT, 'commentaire' TEXT)")
            # enregistrement de la clef primaire
            self.primaryKey = None
            self.cursor.execute(f"PRAGMA table_info({self.tableName})")
            for k in self.cursor.fetchall():
                if k[-1]:
                    self.primaryKey = k[1]
                    self.primaryKeyIndex = k[0]
                    break
            if self.primaryKey is None:
                raise Exit(f"[!] la table '{self.tableName}' de la base de données '{FULLPATH}' n'a pas de clef primaire")
        # le chemin spécifié ne renvois vers rien
        except sqlite3.OperationalError:
            raise Exit(f"[!] la base de donnée '{FULLPATH}' est introuvable") # jamais trigger car connect crée automatiquement un fichier

    # interaction possible avec un 'with'
    def __enter__(self):
        return self
    # interaction possible avec un 'with'
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.save()
        self.close()

    # interaction entre les variables privée et les "getters"
    @property
    def connection(self):
        return self._conn
    @property
    def cursor(self):
        return self._cursor

    # récupere les noms des champs de la table
    def _namesColonnes(self):
        self.cursor.execute(f"PRAGMA table_info({self.tableName})")
        L = [k[1] for k in self.cursor.fetchall()]
        return L

    # renvois True si l'entrée de la clef primaire est bien présente dans la table
    def _verify(self, key, prefixe, suffixe):
        # si prefixe et suffixe valent False, la clef doit exactement etre présente
        if not prefixe and not suffixe:
            self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '{key}'")
        # si seul prefixe vaut True, la clef doit seulement commencer pareil
        elif prefixe and not suffixe:
            self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '{key}%'")
        # si seul suffixe vaut True, la clef doit seulement finir pareil
        elif not prefixe and suffixe:
            self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '%{key}'")
        # si prefixe et suffixe valent True, la clef doit etre contenue
        else:
            self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '%{key}%'")

        # resultat
        if self.cursor.fetchall() == []:
            return False
        else:
            return True

    # recuperer les infos pour une entrée de clef (primaire par défaut) donnée. Si c'est "all", renvoit la totalité des données de la table
    def getDatas(self, username, key, keyname=None, order="date"):
        if not keyname:
            keyname = self.primaryKey
        if key == "all":
            self.cursor.execute(f"SELECT * FROM {self.tableName} WHERE username LIKE '{username}' ORDER BY {order} ASC")
            return self.cursor.fetchall()
        else:
            self.cursor.execute(f"SELECT * FROM {self.tableName} WHERE username LIKE '{username}' AND {keyname} LIKE '{key}'")
            return self.cursor.fetchone()

    # ajoute une nouvelle entrée dans la base de données
    def create(self, valeurs, lower=True):
        nomsColonnes = self._namesColonnes()
        if len(valeurs) != len(nomsColonnes):
            raise Exit(f"[!] les arguments {valeurs} ne correspondent pas au colonnes {nomsColonnes}")
        # on vérifie que l'entrée n'existe pas déja
        if not self._verify(valeurs[self.primaryKeyIndex], False, False):
            text = f"INSERT INTO {self.tableName}("
            for k in nomsColonnes:
                text += f"{k},"
            text = f"{text[:-1]}) VALUES("
            for k in valeurs:
                if k == "NULL":
                    text += "NULL,"
                elif lower:
                    text += f"'{str(k).lower()}',"
                else:
                    text += f"'{k}',"
            text = f"{text[:-1]})"
            try:
                self.cursor.execute(text)
            except sqlite3.OperationalError as e:
                raise Exit(f"[!] erreur dans l'opération : {e}")
        else:
            raise Exit(f"[!] {self.primaryKey} = {valeurs[self.primaryKeyIndex]}, cette entrée existe déjà")

    # supprime une entrée en la selectionnant avec la clef primaire
    def delete(self, key):
        # on vérifie que l'entrée existe
        if self._verify(key, False, False):
            self.cursor.execute(f"DELETE FROM {self.tableName} WHERE {self.primaryKey}='{key}'")
        else:
            raise Exit(f"[!] {self.primaryKey} = {key}, pas d'entrée corespondante")

    # modifie une entrée en la selectionnant avec la clef primaire (dans le champ valeurs)
    def modify(self, valeurs, lower):
        nomsColonnes = self._namesColonnes()
        if len(valeurs) != len(nomsColonnes):
            raise Exit(f"[!] les arguments {valeurs} ne correspondent pas au colonnes {nomsColonnes}")
        # on vérifie que l'entrée existe
        if self._verify(valeurs[self.primaryKeyIndex], False, False):
            text = f"UPDATE {self.tableName} SET"
            for k in range(len(nomsColonnes)):
                if lower:
                    text += f" {nomsColonnes[k]}='{str(valeurs[k]).lower()}',"
                else:
                    text += f" {nomsColonnes[k]}='{valeurs[k]}',"
            text = f"{text[:-1]} WHERE {self.primaryKey} = '{valeurs[self.primaryKeyIndex]}'"
            try:
                self.cursor.execute(text)
            except sqlite3.OperationalError as e:
                raise Exit(f"[!] erreur dans l'opération : {e}")
        else:
            raise Exit(f"[!] {self.primaryKey} = {valeurs[self.primaryKeyIndex]}, pas d'entrée correspondante")

    # sauvegarde la base de donnée
    def save(self):
        self.connection.commit()

    # ferme la base de donnée
    def close(self):
        self.cursor.close()
        self.connection.close()


## ~~~~~~~~~~~~~~~~~~~~~~~~~~    FILTRE MESSAGE PERSO   ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# classe de filtres personalisés
class filtres_perso:
    # detection des dates possibles
    class _date(MessageFilter):
        def filter(self, message):
            try:
                temp = dt.strptime(message.text, "%d %m %Y")
                return True
            except Exception:
                return False
    date = _date()

    # detection des heures possibles
    class _heure(MessageFilter):
        def filter(self, message):
            try:
                temp = dt.strptime(message.text, "%H %M")
                return True
            except Exception:
                return False
    heure = _heure()

    # detection des trajets possibles
    class _trajet(MessageFilter):
        def filter(self, message):
            try:
                return int(message.text) in [0, 1, 2, 3, 4, 5]
            except ValueError:
                return False
    trajet = _trajet()

# renvois sous forme lisible une ligne de la base de donnée
def bdd_to_string(extrait, mode="normal"):
    # structure de la base de donnée
    #  - 0 : id
    #  - 1 : username
    #  - 2 : date
    #  - 3 : date début
    #  - 4 : date fin
    #  - 5 : nb trajets
    #  - 6 : commentaire

    # si mode normal
    if mode == "normal":
        msg = " - le {} de {} à {}".format(
            dt.strptime(extrait[2], '%Y/%m/%d').strftime('%a %-d %B'),
            extrait[3],
            extrait[4],
        )
    # si mode récapitulatif
    elif mode == "recapitulatif":
        msg = "le {} de {} à {} ({} trajets)\ncommentaire: {}".format(
            dt.strptime(extrait[2], '%Y/%m/%d').strftime('%a %-d %B'),
            extrait[3],
            extrait[4],
            extrait[5],
            extrait[6],
        )
    # si mode raccourci
    elif mode == "court":
        msg = f"le {dt.strptime(extrait[2], '%Y/%m/%d').strftime('%a %-d %B')}"
    # si mode affichant seulement la clef primaire
    elif mode == "id":
        msg = extrait[0]

    else:
        print(f"[!] mode inconnu d'affichage : {mode}")

    return msg

## ~~~~~~~~~~~~~~~~~~~~~~~~~~       COMMANDES BOT       ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# fonction lancée par la commande '/start'
def start(update, context):
    update.message.reply_text("Coucou !\nAppuis sur '/' pour voir les commandes disponibles")

# les fonctions de la conversations de nouvelle journee
class conv_nouvelleJournee():
    # conversation de nouvelle journees, commande n°1 de lancement et demande de la date
    def f_new_date(update, context):
        global TO_SAVE
        TO_SAVE = []
        # envois le message de début
        update.message.reply_text(
            "Début de l'enregistrement d'une nouvelle journée\nentre '/stop' pour annuler à tout moment\n\n"
            "quelle est la date (en format 'JJ MM AAAA') de la journees ?",
        )
        # renvoit l'étape suivante
        return DATE

    # conversation de nouvelle journees, commande n°2 pour enregistrer la date et demander l'heure de début
    def f_date_hDebut(update, context):
        global TO_SAVE
        # enregistrement du lieu
        TO_SAVE.append(dt.strptime(update.message.text, "%d %m %Y").strftime("%Y/%m/%d"))
        # la question suivante
        update.message.reply_text("l'heure réelle (en format 'HH MM') de début de journée ?")
        # renvoit l'étape suivante
        return H_DEBUT

    # conversation de nouvelle journees, commande n°3 pour enregistrer l'heure de début et demander l'heure de fin
    def f_hDebut_hFin(update, context):
        global TO_SAVE
        # enregistrement de l'heure de début
        TO_SAVE.append(dt.strptime(update.message.text, "%H %M").strftime("%H:%M"))
        # la question suivante
        update.message.reply_text("l'heure réelle (en format 'HH MM') de fin de journée ?")
        # renvoit l'étape suivante
        return H_FIN

    # conversation de nouvelle journees, commande n°4 pour enregistrer l'heure de fin et demander le nombre de trajet effectué
    def f_hFin_trajet(update, context):
        global TO_SAVE
        # enregistrement de l'heure de fin
        TO_SAVE.append(dt.strptime(update.message.text, "%H %M").strftime("%H:%M"))
        # le clavier qu'on va renvoyer
        keyboard = [["0", "1", "2", "3", "4", "5"]]
        # la question suivante
        update.message.reply_text(
            "combient de trajet on été effectués'heure réelle (en format 'HH MM') de fin de journée ?",
            reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True),
        )
        # renvoit l'étape suivante
        return TRAJET

    # conversation de nouvelle journees, commande n°5 pour enregistrer les trajets et demander un commentaire éventuel
    def f_trajet_commentaire(update, context):
        global TO_SAVE
        # enregistrement du nombre de trajet
        TO_SAVE.append(update.message.text)
        # la question suivante
        update.message.reply_text(
            "un commentaire éventuel ?",
            reply_markup=ReplyKeyboardRemove(),
        )
        # renvoit l'étape suivante
        return COMMENTAIRE

    # conversation de nouvelle journees, commande n°6 pour enregistrer le commentaire et clôturer
    def f_commentaire_sauvegarde(update, context):
        global TO_SAVE
        # enregistrement du commentaire
        TO_SAVE.append(update.message.text)
        # ajout d'un id unique et du nom d'utilisateur
        TO_SAVE = [
            md5(f"{TO_SAVE[0]}_{TO_SAVE[1]}_{update.effective_user.username}".encode()).hexdigest(),
            update.effective_user.username,
        ] + TO_SAVE
        print(TO_SAVE)
        # un petit récapitulatif
        update.message.reply_text(f"récapitulatif :\n{bdd_to_string(TO_SAVE, 'recapitulatif')}")
        # sauvegarde de ces informations dans la base de donnée
        try:
            with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
                temp_bdd.create(TO_SAVE)
            # réponse pour dire que tout va bien
            update.message.reply_text("ok c'est bien enregistré")
            print(f"created : {bdd_to_string(TO_SAVE, 'recapitulatif')}")
        except Exit as e:
            # réponse pour dire qu'il y a eu une erreur
            print("conversation create.f_commentaire_sauvegarde", e) #update.message.reply_text(f"code d'erreur : {e}")

        # remise a zéro des constantes
        TO_SAVE = []
        # fin de la conversation
        return ConversationHandler.END

    # conversation de nouvelle journees, commande d'annulation '/stop'
    def f_stop(update, context):
        global TO_SAVE
        # remise a zéro des constantes
        TO_SAVE = []
        # message d'annulation
        update.message.reply_text(
            "annulation de l'enregistrement",
            reply_markup=ReplyKeyboardRemove()
        )
        # fin de la conversation
        return ConversationHandler.END

# affiche les journees enregistrées dans la base de donnée
def affiche_journees(update, context):
    # toutes les données de la table et tri chronologique
    with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
        temp = temp_bdd.getDatas(update.effective_user.username, "all")
    # si la base de donnée n'est pas vide
    if len(temp) > 0:
        msg = "toutes les journées enregistrées :\n"
        # pour chaque élément on ajoute une ligne
        for k in temp:
            msg += bdd_to_string(k) + "\n"
        update.message.reply_text(msg[:-1])
    # sinon la bdd est vide
    else:
        update.message.reply_text("pas de journee enregistrées :(\nutilises la commande '/nouvelle_journee'")

# supprime une journee de la base de données avec un clavier Inline
def supprime_journee(update, context):
    # le clavuer inline qu'on va remplir
    keyboard = []
    # toutes les données de la table et tri chronologique
    with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
        temp_datas = temp_bdd.getDatas(update.effective_user.username, "all")
    # on les mets dans le lavier inline en colonne
    for k in range(len(temp_datas)):
        keyboard.append([InlineKeyboardButton(bdd_to_string(temp_datas[k], "court"), callback_data="s_"+bdd_to_string(temp_datas[k], "id"))])

    # la ligne pour annuler
    keyboard.append([InlineKeyboardButton("annuler", callback_data="s_annuler")])
    # charge le clavier et l'envois
    update.message.reply_text("sélectionnes pour supprimer :", reply_markup=InlineKeyboardMarkup(keyboard))

# exporte toutes les journées enregistrées dans un fichier excel
def exporte_excel(update, context):
    # nombre de lignes (toutes les données de la table)
    with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
        nb_lignes = len(temp_bdd.getDatas(update.effective_user.username, "all"))
    # si la base de donnée est vide
    if nb_lignes == 0:
        update.message.reply_text("pas de journee enregistrées")
    # sinon on envoit tout
    else:
        # le clavier Inline de confirmation de l'action
        keyboard = [
            [InlineKeyboardButton("envoyer sans nettoyer", callback_data="e_continuer")],
            [InlineKeyboardButton("envoyer et nettoyer ensuite", callback_data="e_continuer_supprimer")],
            [InlineKeyboardButton("annuler", callback_data="e_annuler")],
        ]
        # charge le clavier et l'envois
        update.message.reply_text("quelles options pour créer le fichier Excel ?", reply_markup=InlineKeyboardMarkup(keyboard))

# fonction lancée par un appuis le clavier inline
def button(update, context):
    query = update.callback_query
    # si la query commence par 's', on supprime l'entrée
    if query.data[:2] == "s_":
        query.answer()
        if query.data[2:] == "annuler":
            # réponse au client( change le message précédemment envoyé)
            query.edit_message_text(text="annulé")
        else:
            try:
                with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
                    temp_bdd.delete(query.data[2:])
                # réponse au client (obligatoire sinon bug sur certains clients)
                query.edit_message_text(text="journée supprimée")
                print(f"deleted : {query.data[2:]}")
            except Exit as e:
                # réponse pour dire qu'il y a eu une erreur
                print("fonction button.supprime", e) #query.edit_message_text(text=f"code d'erreur : {e}")

    # si la query commence par 'e', on exporte les journées
    elif query.data[:2] == "e_":
        query.answer()
        # si c'est le code d'annulation
        if query.data[2:] == "annuler":
            # réponse au client( change le message précédemment envoyé)
            query.edit_message_text(text="annulé")

        # si c'est le code de continuation
        if "continuer" in query.data[2:]:
            # toutes les données de la table et tri chronologique
            with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
                temp = temp_bdd.getDatas(update.effective_user.username, "all")
            # nombre de lignes qu'on va écrire
            nb_lignes = len(temp)
            # création du fichier excel temporaire
            wb = Workbook()
            ws = wb.active
            # écriture de la légende
            ws.cell(row=1, column=1).value = "date"
            ws.cell(row=1, column=2).value = "heure début"
            ws.cell(row=1, column=3).value = "heure fin"
            ws.cell(row=1, column=4).value = "nombre trajets"
            ws.cell(row=1, column=5).value = "commentaire"
            # écriture des journées travaillées
            i = 0
            for row in ws.iter_rows(max_col=MAX_COL, min_row=2, max_row=nb_lignes+1):
                # mise en forme compréhensible par excel des données
                row[0].value = dt.strptime(temp[i][2], "%Y/%m/%d").strftime("%d/%m/%Y")
                temp_hdebut = dt.strptime(temp[i][3], "%H:%M")
                row[1].value = (temp_hdebut.hour*60+temp_hdebut.minute)/60
                temp_hfin = dt.strptime(temp[i][4], "%H:%M")
                row[2].value = (temp_hfin.hour*60+temp_hfin.minute)/60
                row[3].value = temp[i][5]
                row[4].value = temp[i][6]
                i += 1
            # le chemin temporaire du excel
            tempPathExcel = os.path.join(BASEPATH, "extrait.xlsx")
            # sauvegarde du excel
            wb.save(filename=tempPathExcel)
            # envoit du fichier
            query.edit_message_text("excel envoyé")
            print("excel envoyé")
            context.bot.send_document(chat_id=query.message.chat_id, document=open(tempPathExcel, "rb"))
            # suppression du excel
            os.remove(tempPathExcel)
            # nettoyage de la base de données si l'ordre est donné
            if "supprimer" in query.data[2:]:
                try:
                    for k in temp:
                        with obj_bdd(BDD_PATH, BDD_TABLE) as temp_bdd:
                            temp_bdd.delete(k[0]) # la clef primaire est en position 0
                            print(f"deleted : {bdd_to_string(k, 'recapitulatif')}")
                            # envoi un nouveau message
                    context.bot.send_message(chat_id=query.message.chat_id, text="base de donnée nettoyée")
                except Exit as e:
                    # réponse pour dire qu'il y a eu une erreur
                    print("fonction button.export", e) #context.bot.send_message(chat_id=query.message.chat_id, text=f"code d'erreur : {e}")

# affiche l'aide
def help(update, context):
    update.message.reply_text("""\
Commandes disponibles:
/nouvelle_journee : enregistre une nouvelle journée de garde
/affiche_journees : affiche toutes les journées
/supprime_journee : supprime une journée
/exporte_excel : renvoit le fichier excel rempli
/help : affiche l'aide""")

# affiche les erreurs rencontrés par le programme
def error(update, context):
    print(f"Update '{update}' \ncaused error '{context.error}'")
    print(format_exc())


## ~~~~~~~~~~~~~~~~~~~~~~~~~~    FONCTION PRINCIPALE    ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# la fonction principale du bot
def main():
    # récupere le token d'identitification dans le .env
    if os.path.isfile(os.path.join(BASEPATH, ".env")):
        with open(os.path.join(BASEPATH, ".env"), "r") as f:
            try:
                # création du bot avec son token d'authentification (retire le 'token=' du début)
                bot = Updater(REGEX_TOKEN.findall(f.read())[0][6:], use_context=True)
            except Exception as e:
                raise e
    else:
        raise Exit("[!] le fichier .env contenant le token d'identitification n'existe pas")
    # initialisation de la base de donnée (crée la base de donnée et la table si elle n'existe pas)
    DATABASE = obj_bdd(BDD_PATH, BDD_TABLE)
    # création du conversation handler pour créer un nouvel enregistrement
    conversation_nouvelleMission = ConversationHandler(
        entry_points=[CommandHandler("nouvelle_journee", conv_nouvelleJournee.f_new_date)],
        states={
            DATE: [MessageHandler(filtres_perso.date, conv_nouvelleJournee.f_date_hDebut)],
            H_DEBUT: [MessageHandler(filtres_perso.heure, conv_nouvelleJournee.f_hDebut_hFin)],
            H_FIN: [MessageHandler(filtres_perso.heure, conv_nouvelleJournee.f_hFin_trajet)],
            TRAJET: [MessageHandler(filtres_perso.trajet, conv_nouvelleJournee.f_trajet_commentaire)],
            COMMENTAIRE: [MessageHandler(Filters.text & ~Filters.command, conv_nouvelleJournee.f_commentaire_sauvegarde)],
        },
        fallbacks=[CommandHandler("stop", conv_nouvelleJournee.f_stop)],
    )
    # ajout des gestionnaires de commande par ordre d'importance
    # la commande /start
    bot.dispatcher.add_handler(CommandHandler("start", start))
    # la commande de conversation /nouvelle_journee
    bot.dispatcher.add_handler(conversation_nouvelleMission)
    # la commande /affiche_journees
    bot.dispatcher.add_handler(CommandHandler("affiche_journees", affiche_journees))
    # la commande /supprime_journee
    bot.dispatcher.add_handler(CommandHandler("supprime_journee", supprime_journee))
    # la commande /exporte_excel
    bot.dispatcher.add_handler(CommandHandler("exporte_excel", exporte_excel))
    # le clavier inline
    bot.dispatcher.add_handler(CallbackQueryHandler(button))
    # la commande /help
    bot.dispatcher.add_handler(CommandHandler("help", help))
    # gestion des erreurs
    bot.dispatcher.add_error_handler(error)

    # lance le bot
    bot.start_polling()
    # continue le programme jusqu'à la reception d'un signal de fin (par ex: CTRL-C)
    bot.idle()

# lance la fonction principale
if __name__ == "__main__":
    main()


# fin
